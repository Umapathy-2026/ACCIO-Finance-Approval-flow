from datetime import datetime, timezone, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_login import login_required, current_user
from werkzeug.security import generate_password_hash
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from app import db
from app.models import User, IssueForm, Ticket, ApprovalLog, Notification, TicketStatus, ApprovalAction, AdminAuditLog
from app.routes.auth import role_required

admin_bp = Blueprint('admin', __name__, url_prefix='/admin', template_folder='../templates/admin')


def log_admin_action(action, target_type=None, target_id=None, details=None):
    entry = AdminAuditLog(
        performed_by=current_user.id,
        action=action,
        target_type=target_type,
        target_id=target_id,
        details=details or {},
        ip_address=request.remote_addr,
        timestamp=datetime.now(timezone.utc)
    )
    db.session.add(entry)


@admin_bp.route('/dashboard')
@login_required
@role_required('admin')
def dashboard():
    now = datetime.now(timezone.utc)
    first_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    total_tickets = Ticket.query.count()
    total_this_month = Ticket.query.filter(Ticket.created_at >= first_of_month).count()
    pending = Ticket.query.filter_by(current_status='Pending').count()
    approved_this_month = Ticket.query.filter(
        Ticket.current_status == 'Sent to Fulfilment',
        Ticket.created_at >= first_of_month
    ).count()
    rejected_this_month = Ticket.query.filter(
        Ticket.current_status == 'Rejected',
        Ticket.created_at >= first_of_month
    ).count()

    recent_tickets = Ticket.query.order_by(Ticket.created_at.desc()).limit(10).all()

    return render_template('admin/dashboard.html',
                         total_tickets=total_tickets,
                         total_this_month=total_this_month,
                         pending=pending,
                         approved_this_month=approved_this_month,
                         rejected_this_month=rejected_this_month,
                         recent_tickets=recent_tickets)


@admin_bp.route('/tickets')
@login_required
@role_required('admin')
def all_tickets():
    tickets = Ticket.query.order_by(Ticket.created_at.desc()).limit(500).all()
    raw_count = Ticket.query.count()
    cap_hit = raw_count > 500
    users = User.query.filter_by(is_active=True).all()
    forms = IssueForm.query.filter_by(is_active=True).all()
    return render_template('admin/all_tickets.html',
                           tickets=tickets, users=users, forms=forms,
                           cap_hit=cap_hit, total=raw_count)


@admin_bp.route('/users', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def users():
    if request.method == 'POST':
        display_name = request.form.get('display_name', '').strip()
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        role = request.form.get('role', 'user')
        manager_id = request.form.get('manager_id', type=int)

        if not display_name or not email or not password:
            flash('Name, email, and password are required.', 'error')
            return redirect(url_for('admin.users'))

        if User.query.filter_by(email=email).first():
            flash('A user with this email already exists.', 'error')
            return redirect(url_for('admin.users'))

        user = User(
            email=email,
            display_name=display_name,
            password_hash=generate_password_hash(password),
            role=role,
            manager_id=manager_id if manager_id else None,
            is_active=True
        )
        db.session.add(user)
        db.session.flush()
        log_admin_action('USER_CREATED', 'user', user.id,
                        {'email': email, 'role': role, 'manager_id': manager_id})
        db.session.commit()
        flash(f'User {display_name} created successfully.', 'success')
        return redirect(url_for('admin.users'))

    users = User.query.order_by(User.created_at.desc()).all()
    approvers = User.query.filter_by(role='approver', is_active=True).all()
    return render_template('admin/users.html', users=users, approvers=approvers, now=datetime.now(timezone.utc))


@admin_bp.route('/users/<int:user_id>/edit', methods=['POST'])
@login_required
@role_required('admin')
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    old_role = user.role

    display_name = request.form.get('display_name', '').strip()
    role = request.form.get('role', 'user')
    manager_id = request.form.get('manager_id', type=int)
    is_active = request.form.get('is_active') == 'on'

    if display_name:
        user.display_name = display_name
    user.role = role
    user.manager_id = manager_id if manager_id else None
    user.is_active = is_active

    password = request.form.get('password', '')
    if password:
        user.password_hash = generate_password_hash(password)

    details = {'role_changed': old_role != role}
    if old_role != role:
        details['from_role'] = old_role
        details['to_role'] = role
    log_admin_action('USER_UPDATED', 'user', user.id, details)
    db.session.commit()
    flash(f'User {user.display_name} updated successfully.', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/users/<int:user_id>/unlock', methods=['POST'])
@login_required
@role_required('admin')
def unlock_user(user_id):
    user = User.query.get_or_404(user_id)
    user.failed_attempts = 0
    user.locked_until = None
    log_admin_action('USER_UNLOCKED', 'user', user.id)
    db.session.commit()
    flash(f'User {user.display_name} has been unlocked.', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/users/<int:user_id>/deactivate-check', methods=['GET'])
@login_required
@role_required('admin')
def deactivate_user_check(user_id):
    user = User.query.get_or_404(user_id)

    open_status_list = [TicketStatus.PENDING.value, TicketStatus.UNDER_REVIEW.value, TicketStatus.NEEDS_CLARIFICATION.value]

    created_open = Ticket.query.filter(
        Ticket.created_by == user_id,
        Ticket.current_status.in_(open_status_list)
    ).all()

    assigned_open = Ticket.query.filter(
        Ticket.assigned_to == user_id,
        Ticket.current_status.in_(open_status_list)
    ).all()

    total_open = len(created_open) + len(assigned_open)
    approvers = User.query.filter_by(role='approver', is_active=True).all()

    return jsonify({
        'has_open_tickets': total_open > 0,
        'open_ticket_count': total_open,
        'created_count': len(created_open),
        'assigned_count': len(assigned_open),
        'created_ticket_ids': [t.id for t in created_open],
        'assigned_ticket_ids': [t.id for t in assigned_open],
        'user_id': user_id,
        'approvers': [{'id': a.id, 'name': a.display_name} for a in approvers]
    })


@admin_bp.route('/users/<int:user_id>/deactivate', methods=['POST'])
@login_required
@role_required('admin')
def deactivate_user(user_id):
    user = User.query.get_or_404(user_id)
    reassign_created_to = request.form.get('reassign_to', type=int)
    reassign_assigned_to = request.form.get('reassign_assigned_to', type=int)

    open_status_list = [TicketStatus.PENDING.value, TicketStatus.UNDER_REVIEW.value, TicketStatus.NEEDS_CLARIFICATION.value]

    created_open = Ticket.query.filter(
        Ticket.created_by == user_id,
        Ticket.current_status.in_(open_status_list)
    ).all()

    assigned_open = Ticket.query.filter(
        Ticket.assigned_to == user_id,
        Ticket.current_status.in_(open_status_list)
    ).all()

    reassigned_count = 0

    if created_open and reassign_created_to:
        new_assignee = User.query.get(reassign_created_to)
        if new_assignee:
            for ticket in created_open:
                ticket.assigned_to = reassign_created_to
                log = ApprovalLog(
                    ticket_id=ticket.id, action_by=current_user.id,
                    action=ApprovalAction.REASSIGNED.value,
                    comment=f'Auto-reassigned from {user.display_name} (deactivated) to {new_assignee.display_name}'
                )
                db.session.add(log)
                reassigned_count += 1

    if assigned_open and reassign_assigned_to:
        new_approver = User.query.get(reassign_assigned_to)
        if new_approver:
            for ticket in assigned_open:
                ticket.assigned_to = reassign_assigned_to
                log = ApprovalLog(
                    ticket_id=ticket.id, action_by=current_user.id,
                    action=ApprovalAction.REASSIGNED.value,
                    comment=f'Approver queue transferred from {user.display_name} (deactivated) to {new_approver.display_name}'
                )
                db.session.add(log)
                reassigned_count += 1

    user.is_active = False
    log_admin_action('USER_DEACTIVATED', 'user', user.id,
                    {'tickets_reassigned': reassigned_count,
                     'created_tickets': len(created_open),
                     'assigned_tickets': len(assigned_open)})
    db.session.commit()

    if reassigned_count > 0:
        flash(f'User deactivated. {reassigned_count} tickets reassigned.', 'success')
    else:
        flash(f'User {user.display_name} deactivated.', 'success')
    return redirect(url_for('admin.users'))


@admin_bp.route('/forms')
@login_required
@role_required('admin')
def forms():
    all_forms = IssueForm.query.order_by(IssueForm.id).all()
    return render_template('admin/forms.html', forms=all_forms)


@admin_bp.route('/forms/<int:form_id>/toggle', methods=['POST'])
@login_required
@role_required('admin')
def toggle_form(form_id):
    form = IssueForm.query.get_or_404(form_id)
    form.is_active = not form.is_active
    log_admin_action('FORM_TOGGLED', 'form', form.id, {'is_active': form.is_active, 'form_name': form.name})
    db.session.commit()
    flash(f'Form "{form.name}" {"activated" if form.is_active else "deactivated"}.', 'success')
    return redirect(url_for('admin.forms'))


@admin_bp.route('/forms/<int:form_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def edit_form(form_id):
    form = IssueForm.query.get_or_404(form_id)

    if request.method == 'POST':
        fields_json = request.form.get('fields_json', '[]')
        import json
        try:
            fields = json.loads(fields_json)
            form.fields = fields
            log_admin_action('FORM_UPDATED', 'form', form.id,
                            {'form_name': form.name, 'field_count': len(fields)})
            db.session.commit()
            flash(f'Form "{form.name}" updated successfully.', 'success')
        except json.JSONDecodeError:
            flash('Invalid fields data.', 'error')

        return redirect(url_for('admin.forms'))

    return render_template('admin/edit_form.html', form=form)


@admin_bp.route('/export')
@login_required
@role_required('admin')
def export_tickets():
    from_dt = request.args.get('from')
    to_dt = request.args.get('to')
    status = request.args.get('status', 'all')
    form_id = request.args.get('form_id', type=int)

    query = Ticket.query

    if from_dt:
        try:
            fd = datetime.strptime(from_dt, '%Y-%m-%d')
            query = query.filter(Ticket.created_at >= fd)
        except: pass
    if to_dt:
        try:
            td = datetime.strptime(to_dt, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Ticket.created_at <= td)
        except: pass
    if status and status != 'all':
        query = query.filter(Ticket.current_status == status)
    if form_id:
        query = query.filter(Ticket.form_id == form_id)

    tickets = query.order_by(Ticket.created_at.desc()).all()

    log_admin_action('EXPORT_TRIGGERED', 'ticket', None,
                    {'from': from_dt, 'to': to_dt, 'status': status, 'count': len(tickets)})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ticket Export'

    header_fill = PatternFill(start_color='01696f', end_color='01696f', fill_type='solid')
    header_font = Font(color='ffffff', bold=True, size=11)
    alt_fill = PatternFill(start_color='f7f6f2', end_color='f7f6f2', fill_type='solid')

    headers = ['Ticket #', 'Issue Type', 'Subject', 'Status', 'Created By', 'Assigned To',
               'Created Date', 'Updated Date', 'Days Open', 'Comments Count']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for i, t in enumerate(tickets, 2):
        row_data = [
            t.ticket_number,
            t.issue_form.name if t.issue_form else '',
            t.subject,
            t.current_status,
            t.creator.display_name if t.creator else '',
            t.assignee.display_name if t.assignee else '',
            t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
            t.updated_at.strftime('%Y-%m-%d %H:%M') if t.updated_at else '',
            (datetime.now(timezone.utc) - t.created_at).days if t.created_at else 0,
            t.approval_logs.count()
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            if (i - 2) % 2 == 1:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = 15
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[col_letter].width = max_len + 2

    filename = f'ACCIO_tickets_{from_dt or "all"}_to_{to_dt or "all"}.xlsx'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/audit-log')
@login_required
@role_required('admin')
def audit_log():
    page = request.args.get('page', 1, type=int)
    logs = AdminAuditLog.query.order_by(
        AdminAuditLog.timestamp.desc()
    ).paginate(page=page, per_page=50, error_out=False)
    return render_template('admin/audit_log.html', logs=logs)