from datetime import datetime, timezone
import os
import io
import uuid
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from sqlalchemy.exc import IntegrityError

from app import db
from app.models import Ticket, IssueForm, ApprovalLog, User, ApprovalAction, TicketStatus, Notification
from app.utils.storage import save_file
from app.utils.email import send_ticket_created, send_clarification_provided
from app.routes.auth import role_required

req_bp = Blueprint('req', __name__, url_prefix='', template_folder='../templates/requester')


def generate_ticket_number():
    import secrets as _secrets
    from datetime import datetime, timezone
    today = datetime.now(timezone.utc).strftime('%Y%m%d')
    for attempt in range(10):
        count = Ticket.query.filter(
            Ticket.ticket_number.like(f'ACC-{today}-%')
        ).count()
        candidate = f'ACC-{today}-{count + 1:04d}'
        if not Ticket.query.filter_by(ticket_number=candidate).first():
            return candidate
    # Fallback: random suffix to avoid collision under high concurrency
    return f'ACC-{today}-{_secrets.token_hex(3).upper()}'


@req_bp.route('/')
def index():
    return redirect(url_for('auth.login'))


@req_bp.route('/dashboard')
@login_required
@role_required('user', 'approver', 'admin')
def dashboard():
    return render_template('requester/dashboard.html')


@req_bp.route('/ticket/new', methods=['GET', 'POST'])
@login_required
@role_required('user', 'approver', 'admin')
def new_ticket():
    if request.method == 'POST':
        if not current_user.manager_id:
            flash(
                'You do not have an approver assigned to your account. '
                'Please contact your administrator before creating tickets.',
                'error'
            )
            return redirect(url_for('req.dashboard'))

        form_id = request.form.get('form_id', type=int)

        if not form_id:
            flash('Please select an issue type.', 'error')
            return redirect(url_for('req.new_ticket'))

        issue_form = IssueForm.query.get_or_404(form_id)

        if not issue_form.is_active:
            flash('This issue type is not available.', 'error')
            return redirect(url_for('req.new_ticket'))

        # Collect dynamic form field values into payload
        payload = {}
        for field in issue_form.fields:
            field_name = field.get('name')
            field_type = field.get('type')
            field_required = field.get('required', False)

            if field_type == 'file':
                file = request.files.get(field_name)
                if file and file.filename:
                    from app.utils.storage import allowed_file
                    is_valid, error_msg = allowed_file(file)
                    if is_valid:
                        original_name, saved_name = save_file(file)
                        payload[field_name] = {
                            'original_name': original_name,
                            'saved_name': saved_name
                        }
                    else:
                        flash(error_msg or f'Invalid file type for {field.get("label")}.', 'error')
                        return redirect(url_for('req.new_ticket'))
            else:
                value = request.form.get(field_name, '').strip()
                if field_required and not value:
                    flash(f'{field.get("label")} is required.', 'error')
                    return redirect(url_for('req.new_ticket'))
                payload[field_name] = value

        # Handle main attachment
        attachment_name = None
        attachment_path = None
        main_file = request.files.get('attachment')
        if main_file and main_file.filename:
            from app.utils.storage import allowed_file
            is_valid, error_msg = allowed_file(main_file)
            if is_valid:
                original_name, saved_name = save_file(main_file)
                attachment_name = original_name
                attachment_path = saved_name

        # Determine assigned approver
        assigned_to = current_user.manager_id
        if not assigned_to:
            approver = User.query.filter_by(role='approver', is_active=True).first()
            if approver:
                assigned_to = approver.id

        # Auto-generate subject from form name if no "subject" field in dynamic fields
        # Also extract description if the form has a "description" field
        subject = None
        description = None
        for field in issue_form.fields:
            fname = field.get('name', '')
            if fname.lower() == 'subject' and payload.get(fname):
                subject = payload.pop(fname)
            if fname.lower() == 'description' and payload.get(fname):
                description = payload.pop(fname)

        if not subject:
            subject = f'{issue_form.name}'

        # Generate ticket number with retry for race conditions
        ticket_number = generate_ticket_number()

        max_retries = 5
        for attempt in range(max_retries):
            ticket = Ticket(
                ticket_number=ticket_number,
                form_id=form_id,
                created_by=current_user.id,
                assigned_to=assigned_to,
                subject=subject,
                description=description,
                payload=payload,
                attachment_name=attachment_name,
                attachment_path=attachment_path,
                current_status='Pending'
            )
            db.session.add(ticket)
            db.session.flush()

            log = ApprovalLog(
                ticket_id=ticket.id,
                action_by=current_user.id,
                action=ApprovalAction.SUBMITTED.value,
                comment='Ticket submitted'
            )
            db.session.add(log)
            try:
                db.session.commit()
                break
            except IntegrityError:
                db.session.rollback()
                if attempt < max_retries - 1:
                    ticket_number = generate_ticket_number()
                    continue
                else:
                    flash('Failed to create ticket due to a concurrency issue. Please try again.', 'error')
                    return redirect(url_for('req.new_ticket'))

        # Notify approver (fail gracefully on Azure free tier)
        if assigned_to:
            approver = User.query.get(assigned_to)
            if approver:
                review_url = url_for('appr.ticket_detail', ticket_id=ticket.id, _external=True)
                try:
                    send_ticket_created(ticket, approver.email, review_url)
                except Exception:
                    import logging
                    logging.exception('Failed to send ticket creation email')

        flash(f'Ticket {ticket_number} created successfully!', 'success')
        return redirect(url_for('req.ticket_detail', ticket_id=ticket.id))

    forms = IssueForm.query.filter_by(is_active=True).all()
    return render_template('requester/new_ticket.html', forms=forms)


@req_bp.route('/ticket/<int:ticket_id>')
@login_required
@role_required('user', 'approver', 'admin')
def ticket_detail(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)

    if current_user.role == 'user' and ticket.created_by != current_user.id:
        flash('You do not have permission to view this ticket.', 'error')
        return redirect(url_for('req.dashboard'))

    logs = ticket.approval_logs.order_by(ApprovalLog.timestamp.desc()).all()
    return render_template('requester/ticket_detail.html', ticket=ticket, logs=logs)


@req_bp.route('/ticket/<int:ticket_id>/clarify', methods=['POST'])
@login_required
@role_required('user', 'approver', 'admin')
def clarify_ticket(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)

    if ticket.created_by != current_user.id:
        flash('You do not have permission to update this ticket.', 'error')
        return redirect(url_for('req.dashboard'))

    if ticket.current_status != 'Needs Clarification':
        flash('This ticket does not need clarification.', 'error')
        return redirect(url_for('req.ticket_detail', ticket_id=ticket.id))

    clarification_text = request.form.get('clarification', '').strip()
    if not clarification_text:
        flash('Please provide clarification information.', 'error')
        return redirect(url_for('req.ticket_detail', ticket_id=ticket.id))

    current_payload = dict(ticket.payload) if ticket.payload else {}
    current_payload['__clarification__'] = clarification_text
    ticket.payload = current_payload

    clar_file = request.files.get('clarification_attachment')
    if clar_file and clar_file.filename:
        from app.utils.storage import allowed_file, save_file
        is_valid, error_msg = allowed_file(clar_file)
        if is_valid:
            original_name, saved_name = save_file(clar_file)
            current_payload['__clarification_attachment__'] = {
                'original_name': original_name,
                'saved_name': saved_name
            }
            ticket.payload = current_payload

    ticket.current_status = TicketStatus.UNDER_REVIEW.value
    ticket.updated_at = datetime.now(timezone.utc)

    log = ApprovalLog(
        ticket_id=ticket.id,
        action_by=current_user.id,
        action=ApprovalAction.CLARIFICATION_PROVIDED.value,
        comment=clarification_text
    )
    db.session.add(log)
    db.session.commit()

    if ticket.assignee:
        review_url = url_for('appr.ticket_detail', ticket_id=ticket.id, _external=True)
        try:
            send_clarification_provided(ticket, ticket.assignee.email, review_url)
        except Exception:
            import logging
            logging.exception('Failed to send clarification notification email')

    flash('Clarification submitted successfully.', 'success')
    return redirect(url_for('req.ticket_detail', ticket_id=ticket.id))


@req_bp.route('/ticket/<int:ticket_id>/export')
@login_required
@role_required('user', 'approver', 'admin')
def export_ticket(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)

    if current_user.role == 'user' and ticket.created_by != current_user.id:
        flash('You do not have permission to export this ticket.', 'error')
        return redirect(url_for('req.dashboard'))

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = 'Ticket Details'

    header_fill = PatternFill(start_color='01696f', end_color='01696f', fill_type='solid')
    header_font = Font(color='ffffff', bold=True, size=11)
    alt_fill = PatternFill(start_color='f7f6f2', end_color='f7f6f2', fill_type='solid')

    details = [
        ('Ticket Number', ticket.ticket_number),
        ('Issue Type', ticket.issue_form.name if ticket.issue_form else ''),
        ('Subject', ticket.subject),
        ('Description', ticket.description or ''),
        ('Status', ticket.current_status),
        ('Created By', ticket.creator.display_name if ticket.creator else ''),
        ('Assigned To', ticket.assignee.display_name if ticket.assignee else ''),
        ('Created At', ticket.created_at.strftime('%Y-%m-%d %H:%M') if ticket.created_at else ''),
        ('Updated At', ticket.updated_at.strftime('%Y-%m-%d %H:%M') if ticket.updated_at else ''),
    ]
    if ticket.payload:
        for key, value in ticket.payload.items():
            if key.startswith('__'):
                continue
            if isinstance(value, dict) and 'original_name' in value:
                value = value['original_name']
            details.append((key.replace('_', ' ').title(), str(value)))

    for col, h in enumerate(['Field', 'Value'], 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for i, (field, val) in enumerate(details, 2):
        ws1.cell(row=i, column=1, value=field).fill = alt_fill if (i-2)%2 else PatternFill()
        ws1.cell(row=i, column=2, value=val).fill = alt_fill if (i-2)%2 else PatternFill()

    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 45

    ws2 = wb.create_sheet('Approval History')
    for col, h in enumerate(['#', 'Action', 'Performed By', 'Comment', 'Timestamp'], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    logs = ticket.approval_logs.order_by(ApprovalLog.timestamp.desc()).all()
    for i, log in enumerate(logs, 2):
        row_data = [i-1, log.action, log.actor.display_name if log.actor else '',
                    log.comment or '',
                    log.timestamp.strftime('%Y-%m-%d %H:%M') if log.timestamp else '']
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            if (i-2) % 2 == 1:
                cell.fill = alt_fill

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 40
    ws2.column_dimensions['E'].width = 20

    filename = f'{ticket.ticket_number}_export.xlsx'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@req_bp.route('/attachment/<int:ticket_id>')
@login_required
def download_attachment(ticket_id):
    """Serve ticket attachment — works for both Azure Blob and local storage."""
    from flask import abort, redirect, send_file
    from app.utils.storage import get_file_path, get_download_url
    import os

    ticket = Ticket.query.get_or_404(ticket_id)

    # Permission: only ticket creator, assigned approver, or admin
    if (current_user.id != ticket.created_by
            and current_user.id != ticket.assigned_to
            and current_user.role != 'admin'):
        abort(403)

    if not ticket.attachment_path:
        abort(404)

    # Azure Blob: redirect to time-limited SAS URL
    sas_url = get_download_url(ticket.attachment_path, ticket.attachment_name)
    if sas_url:
        return redirect(sas_url)

    # Local dev fallback
    filepath = get_file_path(ticket.attachment_path)
    if not os.path.exists(filepath):
        abort(404)
    return send_file(
        filepath,
        as_attachment=True,
        download_name=ticket.attachment_name or ticket.attachment_path
    )


@req_bp.route('/notifications')
@login_required
def notifications():
    all_notifs = Notification.query.filter_by(user_id=current_user.id)\
        .order_by(Notification.created_at.desc()).all()
    Notification.query.filter_by(user_id=current_user.id, is_read=False)\
        .update({'is_read': True})
    db.session.commit()
    return render_template('requester/notifications.html', notifications=all_notifs)